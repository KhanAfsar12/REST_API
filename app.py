from flask import Flask, jsonify, request
from flask_restful import Resource, Api
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook

app = Flask(__name__)
api = Api(app)

app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:root@localhost:5432/afsar'
db = SQLAlchemy(app)



class HelloWorld(Resource):
    def get(self):
        return {'data' : 'Hello World!'}
    
class HelloName(Resource):
    def get(self, Name):
        return {'data' : 'Hello {}'.format(Name)}

class Integer(Resource):
    def get(self, num):
        if num%2==0:
            return "Even: {}".format(num)
        else:
            return "odd: {}".format(num)

api.add_resource(HelloWorld, '/helloworld')
api.add_resource(HelloName, '/helloworld/<string:Name>')
api.add_resource(Integer, '/TODO/<int:num>')


# It is a model
class Crud(db.Model):
    id =db.Column(db.Integer, primary_key=True)
    mark = db.Column(db.String(200))
    rank = db.Column(db.String(200))
    year = db.Column(db.Integer)
    is_active = db.Column(db.String(200))


# Inserting data into database via postman with the help of following code
@app.route('/gear', methods=['POST'])
def func():
    if request.method == "POST":
        f = request.files['book']
        wb = load_workbook(f)
        a = wb.active

        for i in a.iter_rows(min_row=2, values_only=True):
            print("-------------------------------ITERATED------------------------------------")
            cd = Crud( mark=str(i[0]), rank=str(i[1]), year=i[2], is_active=i[3])
            print("&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&  OBJ CREATED  &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&")
            db.session.add(cd)
            print("?????????????????????????????????????  added  ???????????????????????????????????????")
        db.session.commit()
        print("@@@@@@@@@@@@@@@@@@@@@@@@@  Commit  @@@@@@@@@@@@@@@@@@@@@@@@@@@@")
    return "Afsar Khan"


# Creating API for retrieval operation
@app.route('/ret', methods=['POST'])
def get_data():
    access = Crud.query.all()
    output = []
    for i in access:
            access_list ={'id': i.id, 'mark':i.mark, 'rank':i.rank, 'year':i.year, 'is_active':i.is_active}
            output.append(access_list)    
    return {"i" : output}


# It is a retrieval operation by id
@app.route('/ret/<id>')
def get_num(id):
    attempt = Crud.query.get_or_404(id)
    return {'mark' : attempt.mark, 'rank': attempt.rank, 'year' : attempt.year, 'is_active' : attempt.is_active}
    

# Creating API for insertion
@app.route('/insert', methods=['POST'])
def set_num():
    insert = Crud(mark=request.json['mark'], rank=request.json['mark'], year=request.json['year'], is_active=request.json['is_active'])
    db.session.add(insert)
    db.session.commit()
    return {'id' : insert.id}


# Creating API for Deletion
@app.route('/del/<id>', methods=['DELETE'])
def dear(id):
    map = Crud.query.get(id)
    if map is None:
        return {'error' : 'NotFound Value'}
    db.session.delete(map)
    db.session.commit()
    return {"message" : "Deleted"}


if __name__ == '__main__':
    app.run(debug=True, port=6000)
